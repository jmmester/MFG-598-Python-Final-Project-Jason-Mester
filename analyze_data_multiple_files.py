# -*- coding: utf-8 -*-
"""
Created on Fri Apr 21 11:23:29 2023

@author: Jason
"""
#This script analyzes the data found in the other scripts
#It determines whether the topic is plausible
#This script will analyze the data for any amount of files
import pandas as pd
from openpyxl.styles import Font, colors, PatternFill, Border, Side
import os

#loop over the numbers 1 to 5 and get the CSV file of that number
#If you had more files then you could add your amount to where the 6 
for i in range(1, 6):
    filename = f'amazon_book_info_search_page_{i}.csv'
    if not os.path.isfile(filename):
        # if the CSV file does not exist skip to the next iteration
        continue

    # read CSV
    df = pd.read_csv(filename)

    # creat a column and if Reviews is greater than 150 and revenue is greater than 500 then it is plausible so it wold be a yes
    df['Plausible'] = 'No'
    df.loc[(df['Reviews'] > 150) & (df['Monthly revenue'] > 500), 'Plausible'] = 'Yes'

    # create an Excel based off of the CSV so that way the formatting can be made
    excel_filename = f'amazon_book_info_search_page_{i}ex.xlsx'
    writer = pd.ExcelWriter(excel_filename, engine='openpyxl')
    df.to_excel(writer, index=False)

    # get the workbook and worksheet objects
    workbook = writer.book
    worksheet = writer.sheets['Sheet1']

    # set the width of columns A, B and C
    #Certain width is set for each column so that way it is easy to read when it opens
    worksheet.column_dimensions['A'].width = 75.45
    worksheet.column_dimensions['B'].width = 7
    worksheet.column_dimensions['C'].width = 5
    worksheet.column_dimensions['D'].width = 20
    worksheet.column_dimensions['E'].width = 6
    worksheet.column_dimensions['F'].width = 13
    worksheet.column_dimensions['G'].width = 14.5
    worksheet.column_dimensions['H'].width = 7.5

    # define the cell styles for 'Yes' and 'No'
    #If it is a yes then the row turns green and if it is a no the row turns red
    #It also puts a border around every cell to make it look more like a chart
    font_yes = Font(color=colors.BLACK)
    fill_yes = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
    font_no = Font(color=colors.WHITE)
    fill_no = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000'))

    #This goes thorugh each row and applies the formatting depending on the plausible column
    for j, row in df.iterrows():
        if row['Plausible'] == 'Yes':
            fill = fill_yes
            font = font_yes
        else:
            fill = fill_no
            font = font_no
        for k, value in enumerate(row):
            cell = worksheet.cell(row=j+2, column=k+1)
            cell.fill = fill
            cell.font = font
            cell.border = border

    # save the Excel file
    workbook.save(f'amazon_book_info_search_page_{i}ex.xlsx')

    # open the Excel file to view the sesults
    os.startfile(f'amazon_book_info_search_page_{i}ex.xlsx')
